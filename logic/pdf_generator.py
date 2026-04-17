"""
買付証明書PDF生成パイプライン（Cowork実装統合版）

流程：
1. Excelテンプレート + データ → XLSM
2. XLSM → PDF (LibreOffice UNO経由)
3. XLSM から EMF社印 を抽出 → Ghostscript で PNG化
4. PDF + PNG社印 → オーバーレイ合成
"""

import zipfile
import subprocess
import time
import os
import io
import struct
import logging
from pathlib import Path
from typing import Optional
from utils.config import config

logger = logging.getLogger(__name__)

try:
    from PIL import Image
    import numpy as np
    from reportlab.pdfgen import canvas
    from reportlab.lib.utils import ImageReader
    import pypdf
except ImportError as e:
    print(f"Warning: Missing image processing library: {e}")


class PDFGenerator:
    """Excel → PDF + 社印合成（Cowork実装）"""

    def __init__(self, excel_path: str, output_path: str):
        self.excel_path = Path(excel_path)
        self.output_path = Path(output_path)
        self.temp_dir = Path(output_path).parent / 'temp'
        self.temp_dir.mkdir(parents=True, exist_ok=True)

    def generate_with_seal(self) -> bool:
        """
        XLSM → PDF（社印入り）生成
        流程：
        1. LibreOffice UNO → PDF (社印なし)
        2. EMF → PNG (Ghostscript)
        3. PDF + PNG → オーバーレイ合成
        """
        try:
            # Step 1: LibreOffice で XLSM → PDF
            temp_pdf_noseal = self.temp_dir / 'output_noseal.pdf'
            if not self._export_pdf_via_libreoffice(str(self.excel_path), str(temp_pdf_noseal)):
                return False

            # Step 2: EMF → PNG (社印)
            seal_png = self.temp_dir / 'seal.png'
            if not self._extract_seal_from_xlsm(str(self.excel_path), str(seal_png)):
                # EMF抽出失敗時は社印なしで返す
                temp_pdf_noseal.rename(self.output_path)
                return True

            # Step 3: PDF + 社印 → オーバーレイ合成
            if not self._overlay_seal_on_pdf(str(temp_pdf_noseal), str(seal_png), str(self.output_path)):
                # 合成失敗時は社印なしで返す
                temp_pdf_noseal.rename(self.output_path)
                return True

            return True

        except Exception as e:
            logger.error(f"Error generating PDF: {e}", exc_info=True)
            return False

    @staticmethod
    def _export_pdf_via_libreoffice(xlsm_path: str, pdf_path: str) -> bool:
        """
        LibreOffice ヘッドレスコマンドでXLSM→PDF変換
        「買付証明書」シート（Sheet 1）のみをPDF化
        """
        try:
            import openpyxl
            import shutil

            output_dir = str(Path(pdf_path).parent)
            temp_xlsx = Path(output_dir) / 'temp_kaituke_only.xlsx'

            # 元のファイルをコピー
            shutil.copy(xlsm_path, str(temp_xlsx))

            # Excelを処理
            wb = openpyxl.load_workbook(str(temp_xlsx), data_only=False)
            ws_target = wb['買付証明書']

            # A30 に定型文を入力（数式を削除して置き換え）
            ws_target['A30'].value = '　素敵な物件にご縁をいただき感謝申し上げます。'

            # A31:A34 は空白にする
            for row_num in range(31, 35):
                ws_target[f'A{row_num}'].value = None

            # 「買付証明書」をアクティブシートに設定
            if '買付証明書' in wb.sheetnames:
                wb.active = wb.index(wb['買付証明書'])

            # 「入力用」シートを削除
            if '入力用' in wb.sheetnames:
                del wb['入力用']

            wb.save(str(temp_xlsx))
            wb.close()

            # LibreOfficeでPDF化
            cmd = [
                config.LIBREOFFICE_PATH,
                '--headless',
                '--convert-to', 'pdf',
                '--outdir', output_dir,
                str(temp_xlsx)
            ]

            result = subprocess.run(cmd, capture_output=True, timeout=60, check=False)

            # LibreOfficeは元のファイル名.pdfを生成するため、リネーム
            generated_pdf = Path(output_dir) / 'temp_kaituke_only.pdf'
            if generated_pdf.exists():
                generated_pdf.rename(pdf_path)
                if temp_xlsx.exists():
                    temp_xlsx.unlink()  # 一時ファイル削除
                return True
            else:
                logger.error(f"Error: Generated PDF not found at {generated_pdf}")
                logger.error(f"stderr: {result.stderr.decode()}")
                if temp_xlsx.exists():
                    temp_xlsx.unlink()
                return False

        except Exception as e:
            logger.error(f"Error in LibreOffice PDF export: {e}", exc_info=True)
            return False

    @staticmethod
    def _extract_seal_from_xlsm(xlsm_path: str, seal_png_path: str) -> bool:
        """
        XLSM → EMF社印 → PNG (Ghostscript)

        EMF形式は GDICコメント (offset 19056) 内に埋め込まれたPDFを含む。
        このPDFを抽出してGhostscriptで描画。
        """
        try:
            with zipfile.ZipFile(xlsm_path, 'r') as z:
                emf_data = z.read('xl/media/image1.emf')

            # EMF のGDICコメント（offset 19056）から PDF を抽出
            offset = 19056
            data_size = struct.unpack('<I', emf_data[offset+8:offset+12])[0]
            gdic_data = emf_data[offset+12:offset+12+data_size]

            pdf_start = gdic_data.find(b'%PDF')
            eof_pos = gdic_data.rfind(b'%%EOF')
            seal_pdf_bytes = gdic_data[pdf_start:eof_pos+5]

            # 一時PDFに書き出す
            tmp_pdf = seal_png_path.replace('.png', '_tmp.pdf')
            with open(tmp_pdf, 'wb') as f:
                f.write(seal_pdf_bytes)

            # Ghostscript で PNG に変換
            gs_path = config.GHOSTSCRIPT_PATH
            result = subprocess.run([
                gs_path, '-dBATCH', '-dNOPAUSE', '-sDEVICE=png256',
                '-r300', f'-sOutputFile={seal_png_path}', tmp_pdf
            ], check=False, capture_output=True, timeout=30)

            os.remove(tmp_pdf)
            return Path(seal_png_path).exists()

        except Exception as e:
            logger.error(f"Error extracting seal: {e}", exc_info=True)
            return False

    @staticmethod
    def _overlay_seal_on_pdf(cert_pdf: str, seal_png: str, output_pdf: str) -> bool:
        """
        証明書PDF に社印PNG をオーバーレイ合成

        社印配置位置（Excel drawing2.xml から抽出）:
        - セル AC6 〜 AH11
        - 正確な座標: X=243.4pt, Y=678.2pt（左上基準）
        - サイズ: 77.3pt × 72.8pt
        - 回転: 86.8° CCW
        """
        try:
            PAGE_W, PAGE_H = 595.27, 841.89   # A4 size in points
            # 右上の「6とDの横」に配置（売主情報の右側）
            SEAL_X = 454.0                     # pts（3分の1左に移動 = 480 - 77.3/3）
            SEAL_Y_TOP = 175.0                 # pts（少し下）
            SEAL_W = 77.3                      # pts（幅）
            SEAL_H = 72.8                      # pts（高さ）
            SEAL_ROT = 0                       # degrees（回転なし）

            # 白を透過に変換
            img = Image.open(seal_png).convert("RGBA")
            arr = np.array(img)
            white = (arr[:, :, 0] > 240) & (arr[:, :, 1] > 240) & (arr[:, :, 2] > 240)
            arr[white, 3] = 0
            img = Image.fromarray(arr)

            # ReportLab でオーバーレイ層を作成
            img_buf = io.BytesIO()
            img.save(img_buf, format='PNG')
            img_buf.seek(0)

            overlay_buf = io.BytesIO()
            c = canvas.Canvas(overlay_buf, pagesize=(PAGE_W, PAGE_H))
            cx = SEAL_X + SEAL_W / 2
            cy = PAGE_H - SEAL_Y_TOP - SEAL_H / 2
            c.saveState()
            c.translate(cx, cy)
            c.rotate(SEAL_ROT)
            c.drawImage(ImageReader(img_buf), -SEAL_W/2, -SEAL_H/2,
                        width=SEAL_W, height=SEAL_H, mask='auto')
            c.restoreState()
            c.save()
            overlay_buf.seek(0)

            # pypdf で合成
            reader_cert = pypdf.PdfReader(cert_pdf)
            reader_overlay = pypdf.PdfReader(overlay_buf)
            writer = pypdf.PdfWriter()
            page = reader_cert.pages[0]
            page.merge_page(reader_overlay.pages[0])
            writer.add_page(page)

            with open(output_pdf, "wb") as f:
                writer.write(f)

            return True

        except Exception as e:
            logger.error(f"Error overlaying seal: {e}", exc_info=True)
            return False

    def convert_pdf_to_image(self) -> Optional[str]:
        """
        生成されたPDFをA4サイズの画像に変換
        返り値: 画像ファイルパス（成功時）/ None（失敗時）
        """
        try:
            from pdf2image import convert_from_path

            if not self.output_path.exists():
                logger.error(f"PDF file not found: {self.output_path}")
                return None

            # PDFを画像に変換（150 dpi で十分）
            images = convert_from_path(str(self.output_path), dpi=150, first_page=1, last_page=1)

            if not images:
                logger.error("Failed to convert PDF to image")
                return None

            # 画像をJPEGで保存（LINEの送信に最適）
            image_path = str(self.output_path).replace('.pdf', '.jpg')
            images[0].save(image_path, 'JPEG', quality=95)

            logger.info(f"PDF converted to image: {image_path}")
            return image_path

        except Exception as e:
            logger.error(f"Error converting PDF to image: {e}", exc_info=True)
            return None

    def cleanup(self):
        """一時ファイルをクリーンアップ"""
        try:
            import shutil
            if self.temp_dir.exists():
                shutil.rmtree(self.temp_dir)
        except Exception as e:
            logger.warning(f"Error cleaning up: {e}")
